import os
import sys
from _csv import reader

import openpyxl
## Import Pyside
#from PySide2.QtGui import QPainter
#from PySide2.QtCharts import QtCharts
#from PySide2.QtCore import *

from PyQt5.QtChart import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *

## import Math and MOD
from random import randrange
from functools import partial

import csv


## import GUI file
from thomasights import *

## import custom widget
from Custom_Widgets.Widgets import *

## object_tracker (Basic) Imports
from absl import flags
from datetime import datetime
FLAGS = flags.FLAGS
FLAGS(sys.argv)

import time
import numpy as np
import cv2
import matplotlib.pyplot as plt
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

import tensorflow as tf
from yolov3_tf2.models import YoloV3
from yolov3_tf2.dataset import transform_images
from yolov3_tf2.utils import convert_boxes

from deep_sort import preprocessing
from deep_sort import nn_matching
from deep_sort.detection import Detection
from deep_sort.tracker import Tracker
from tools import generate_detections as gdet


class CustomSlider(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setRange(0, 100)  # Adjust this as needed

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        style = self.style()
        opt = QStyleOptionSlider()
        opt.initFrom(self)
        opt.subControls = QStyle.SC_Slider | QStyle.SC_SliderHandle
        style.drawControl(QStyle.CE_Slider, opt, painter)

        # Draw the triangular handle overlay
        painter.setBrush(QColor("red"))  # Change this to your desired color
        painter.translate(self.slider.sliderPosition(), 0)
        painter.rotate(45)
        triangle_path = QPainterPath()
        triangle_path.moveTo(0, -7)
        triangle_path.lineTo(7, 7)
        triangle_path.lineTo(-7, 7)
        triangle_path.close()
        painter.drawPath(triangle_path)


class VideoProcessor(QThread):

    frame_processed = pyqtSignal(np.ndarray)
    current_activity_changed = pyqtSignal(str)


    def __init__(self, parent=None):
        super(VideoProcessor, self).__init__(parent)
        self.video_path = None
        self.cap = None
        self.current_activity = None
        self.ThreadActive = False
        self.moveToThread(QThread.currentThread())  # Move to the main thread initially

    def set_current_activity(self, activity):
        # Set the current_activity in VideoProcessor
        self.current_activity = activity
        print("Chosen Activity in VideoProcessor:", self.current_activity)

    def run(self):
        self.ThreadActive = True
        self.moveToThread(self.thread())  # Move to its own thread
        if self.video_path is not None:
            self.cap = cv2.VideoCapture(self.video_path)
            if not self.cap.isOpened():
                print("Error: Could not open video.")
                return

            # Initialize your object detector and tracker here
            # ...

            class_names = [c.strip() for c in open('../data/labels/coco.names').readlines()]
            yolo = YoloV3(classes=len(class_names))
            yolo.load_weights('../weights/yolov3.tf')

            max_cosine_distance = 0.5
            nn_budget = None
            nms_max_overlap = 0.8

            model_filename = '../model_data/mars-small128.pb'
            encoder = gdet.create_box_encoder(model_filename, batch_size=1)
            metric = nn_matching.NearestNeighborDistanceMetric('cosine', max_cosine_distance, nn_budget)
            tracker = Tracker(metric)

            codec = cv2.VideoWriter_fourcc(*'XVID')
            vid_fps = int(self.cap.get(cv2.CAP_PROP_FPS))
            vid_width, vid_height = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(
                self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            out = cv2.VideoWriter('../data/video/results.avi', codec, vid_fps, (vid_width, vid_height))

            class_counts = {
                'clapping': 0,
                'tapping': 0,
                'bored': 0,
                'listening': 0,
                'curious': 0,
                'surprised': 0,
                'writing': 0,
                'leaning': 0,
                'sad': 0,
                'confused': 0,
                'focused': 0,
                'angry': 0,
                'scared': 0,
                'happy': 0,
            }

            activity_value = str(self.current_activity)

            shades_of_purple_rgb = [
                (83, 91, 191),  # #535BBF Semi Deep Purple
                (122, 130, 233),  # #7A82E9 Lavander
                (71, 76, 138),  # #474C8A Deep Purple
                (74, 82, 191),  # #4A52BF Iconic Light Purple
                (140, 82, 255)  # #8C52FF Mmmmm Purple
            ]

            # Convert RGB to BGR
            shades_of_purple_bgr = [(color[2], color[1], color[0]) for color in shades_of_purple_rgb]

            # Excel and CSV File for Database
            excel_file_path = '../data/database/test2.xlsx'
            csv_file_path = '../data/database/test2.csv'

            # Set to keep track of unique detection IDs
            unique_detection_ids = set()

            # Outside the loop
            if os.path.exists(excel_file_path):
                historical_data = pd.read_excel(excel_file_path)
            else:
                # If the file doesn't exist, create an empty DataFrame
                historical_data = pd.DataFrame(columns=[' date ', '  time '] + list(class_counts.keys()))

            # Historical Trajectory and Countings
            from _collections import deque
            pts = [deque(maxlen=30) for _ in range(1000)]

            while self.ThreadActive:
                ret, img = self.cap.read()
                if not ret:
                    print('Completed')
                    break

                # Your video processing logic here
                # ...

                img_in = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                img_in = tf.expand_dims(img_in, 0)
                img_in = transform_images(img_in, 416)
                t1 = time.time()

                boxes, scores, classes, nums = yolo.predict(img_in)

                classes = classes[0]
                names = []
                for i in range(len(classes)):
                    names.append(class_names[int(classes[i])])
                names = np.array(names)
                converted_boxes = convert_boxes(img, boxes[0])
                features = encoder(img, converted_boxes)

                detections = [Detection(bbox, score, class_name, feature) for bbox, score, class_name, feature in
                              zip(converted_boxes, scores[0], names, features)]

                boxs = np.array([d.tlwh for d in detections])
                scores = np.array([d.confidence for d in detections])
                classes = np.array([d.class_name for d in detections])
                indices = preprocessing.non_max_suppression(boxs, classes, nms_max_overlap, scores)
                detections = [detections[i] for i in indices]

                tracker.predict()
                tracker.update(detections)

                cmap = plt.get_cmap('tab20b')
                colors = [cmap(i)[:3] for i in np.linspace(0, 1, 20)]

                current_count = int(0)

                for track in tracker.tracks:
                    if not track.is_confirmed() or track.time_since_update > 1:
                        continue
                    bbox = track.to_tlbr()
                    class_name = track.get_class()
                    shade_index = int(track.track_id) % len(shades_of_purple_bgr)
                    color = shades_of_purple_bgr[shade_index]

                    cv2.rectangle(img, (int(bbox[0]), int(bbox[1])), (int(bbox[2]), int(bbox[3])), color, 2)
                    cv2.rectangle(img, (int(bbox[0]), int(bbox[1] - 30)),
                                  (int(bbox[0]) + (len(class_name) + len(str(track.track_id))) * 17, int(bbox[1])),
                                  color, -1)
                    cv2.putText(img, class_name + "-" + str(track.track_id), (int(bbox[0]), int(bbox[1] - 10)), 0, 0.75,
                                (255, 255, 255), 2)

                    # In this line includes how to incorporate Hystorical Trajectory in MOT
                    # Center Point
                    center = (int(((bbox[0]) + (bbox[2])) / 2), int(((bbox[1]) + (bbox[3])) / 2))
                    pts[track.track_id].append(center)

                    for j in range(1, len(pts[track.track_id])):
                        if pts[track.track_id][j - 1] is None or pts[track.track_id][j] is None:
                            continue
                        thickness = int(np.sqrt(64 / float(j + 1)) * 2)
                        cv2.line(img, (pts[track.track_id][j - 1]), (pts[track.track_id][j]), color, thickness)

                    class_name_lower = class_name.lower()
                    if class_name_lower in class_counts and track.track_id not in unique_detection_ids:
                        class_counts[class_name_lower] += 1
                        unique_detection_ids.add(track.track_id)

                    print("Class Counts:", class_counts)
                    print("Detection ID:", track.track_id)
                    print("Number of Detections:", nums)



                fps = 1. / (time.time() - t1)
                cv2.putText(img, "FPS: {:.2f}".format(fps), (0, 30), 0, 1, (255, 255, 255), 2)
                out.write(img)

                # Emit the processed frame to update GUI
                self.frame_processed.emit(img)

                if cv2.waitKey(1) == ord('q'):
                    break


            # Save the final counts to the Excel file outside the loop
            if nums > 0:  # Only update if there are detections

                current_datetime = datetime.now()
                data = {
                    'date': current_datetime.strftime("%Y-%m-%d"),
                    'time': current_datetime.strftime("%H:%M:%S"),
                    'activity': activity_value,
                }
                data.update(class_counts)

                historical_data = historical_data.append(data, ignore_index=True)

                # Reset index before saving to Excel
                historical_data.reset_index(drop=True, inplace=True)

                book = load_workbook(excel_file_path)
                writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                historical_data.to_excel(writer, index=False, sheet_name='Sheet1')
                writer.save()

                # Save to CSV
                historical_data.to_csv(csv_file_path, index=False)

            # Release VideoCapture after processing
            self.cap.release()
            cv2.destroyAllWindows()

    def stop(self):
        self.ThreadActive = False

class MainWindow(QMainWindow):

    current_activity = pyqtSignal(str)

    def __init__(self, parent=None):

        self.main_win = QMainWindow()
        QMainWindow.__init__(self)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # Apply JSON Stylesheet
        loadJsonStyle(self, self.ui)

        # Button Navigations

        self.ui.stackedWidget.setCurrentWidget(self.ui.detection)
        self.ui.detection_btn.clicked.connect(self.showDetection)
        self.ui.bar_graph_btn.clicked.connect(self.showBarGraph)
        self.ui.nested_donuts_btn.clicked.connect(self.showNestedDonuts)
        self.ui.donutback_btn1.clicked.connect(self.showNestedDonuts)
        self.ui.donutnext_btn1.clicked.connect(self.showNestedDonuts1)
        self.ui.donutback_btn2.clicked.connect(self.showNestedDonuts1)
        self.ui.donutnext_btn2.clicked.connect(self.showNestedDonuts2)
        self.ui.line_chart_btn.clicked.connect(self.showLineChart)
        self.ui.record_btn.clicked.connect(self.showHybridGraph)
        self.ui.prediction_btn.clicked.connect(self.showPrediction)
        self.ui.auditoryPushButton.clicked.connect(self.showAuditorySuggestions)
        self.ui.auditorySuggest_btn1.clicked.connect(self.showAuditorySuggestions1)
        self.ui.auditorySuggest_btn2.clicked.connect(self.showAuditorySuggestions2)
        self.ui.auditorySuggest_btn3.clicked.connect(self.showAuditorySuggestions3)
        self.ui.auditorySuggest_btn4.clicked.connect(self.showAuditorySuggestions4)
        self.ui.visualPushButton.clicked.connect(self.showVisualSuggestions)
        self.ui.visualSuggest_btn1.clicked.connect(self.showVisualSuggestions1)
        self.ui.visualSuggest_btn2.clicked.connect(self.showVisualSuggestions2)
        self.ui.visualSuggest_btn3.clicked.connect(self.showVisualSuggestions3)
        self.ui.visualSuggest_btn4.clicked.connect(self.showVisualSuggestions4)
        self.ui.kinestheticPushButton.clicked.connect(self.showKinestheticSuggestions)
        self.ui.kinestheticSuggest_btn1.clicked.connect(self.showKinestheticSuggestions1)
        self.ui.kinestheticSuggest_btn2.clicked.connect(self.showKinestheticSuggestions2)
        self.ui.kinestheticSuggest_btn3.clicked.connect(self.showKinestheticSuggestions3)
        self.ui.kinestheticSuggest_btn4.clicked.connect(self.showKinestheticSuggestions4)

        # Back Buttons

        self.ui.audioSuggest_Backbtn1.clicked.connect(self.showAuditorySuggestions)
        self.ui.audioSuggest_Backbtn2.clicked.connect(self.showAuditorySuggestions)
        self.ui.audioSuggest_Backbtn3.clicked.connect(self.showAuditorySuggestions)
        self.ui.audioSuggest_Backbtn4.clicked.connect(self.showAuditorySuggestions)
        self.ui.visualSuggest_Backbtn1.clicked.connect(self.showVisualSuggestions)
        self.ui.visualSuggest_Backbtn2.clicked.connect(self.showVisualSuggestions)
        self.ui.visualSuggest_Backbtn3.clicked.connect(self.showVisualSuggestions)
        self.ui.visualSuggest_Backbtn4.clicked.connect(self.showVisualSuggestions)
        self.ui.kinestheticSuggest_Backbtn1.clicked.connect(self.showKinestheticSuggestions)
        self.ui.kinestheticSuggest_Backbtn2.clicked.connect(self.showKinestheticSuggestions)
        self.ui.kinestheticSuggest_Backbtn3.clicked.connect(self.showKinestheticSuggestions)
        self.ui.kinestheticSuggest_Backbtn4.clicked.connect(self.showKinestheticSuggestions)

# Set Window Minimum Size
        self.setMinimumSize(950, 700)

# Show Window
        self.show()
        self.create_bar() # Show Bar Series
        self.create_donutchart() # Show Donut Chart
        self.create_donutchart1()
        self.create_donutchart2()
        self.create_linegraph()

# Play Gifs
        self.alignCenterPercentage()
        self.playAuditoryGif()
        self.playVisualGif()
        self.playKinestheticGif()
        self.playAuditorySuggestGif1()
        self.playAuditorySuggestGif2()
        self.playAuditorySuggestGif3()
        self.playAuditorySuggestGif4()
        self.playVisualSuggestGif1()
        self.playVisualSuggestGif2()
        self.playVisualSuggestGif3()
        self.playVisualSuggestGif4()
        self.playKinestheticSuggestGif1()
        self.playKinestheticSuggestGif2()
        self.playKinestheticSuggestGif3()
        self.playKinestheticSuggestGif4()

# Fix Layouts
        self.fixAuditorySuggestions1()
        self.fixAuditorySuggestions2()
        self.fixAuditorySuggestions3()
        self.fixAuditorySuggestions4()
        self.fixVisualSuggestions1()
        self.fixVisualSuggestions2()
        self.fixVisualSuggestions3()
        self.fixVisualSuggestions4()
        self.fixKinestheticSuggestions1()
        self.fixKinestheticSuggestions2()
        self.fixKinestheticSuggestions3()
        self.fixKinestheticSuggestions4()


        self.video_processor = VideoProcessor()
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.video_processor.frame_processed.connect(self.update_gui)

        self.ui.comboBox.currentIndexChanged.connect(self.update_activity)
        self.current_activity.connect(self.video_processor.set_current_activity)

        # Manually call update_activity to set the initial current_activity
        self.update_activity(self.ui.comboBox.currentIndex())


        # Start VideoProcessor in its own thread
        self.video_processor_thread = QThread(self)
        self.video_processor.moveToThread(self.video_processor_thread)
        self.video_processor_thread.started.connect(self.video_processor.run)
        self.video_processor_thread.start()

        # Connect button click to video upload function
        self.ui.upload_detectionBtn.clicked.connect(self.upload_video)
        self.ui.live_detectionBtn.clicked.connect(self.real_time)


        # Date and Time Timer
        # Create a timer that triggers the update every second
        self.dtimer = QTimer(self)
        self.dtimer.timeout.connect(self.update_time)
        self.dtimer.start(1000)  # Update every 1000 milliseconds (1 second)

        self.ui.exit_detectionBtn.clicked.connect(self.stop_detection)
        self.ui.exit_detectionBtn.clicked.connect(self.refresh_table)
        self.ui.exit_detectionBtn.clicked.connect(self.success_message)

        self.custom_slider = CustomSlider()
        self.layout

        # Initial update
        self.update_time()
        self.load_records()


    def upload_video(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_dialog = QFileDialog()
        file_dialog.setNameFilter("Video Files (*.mp4 *.avi)")
        file_dialog.setWindowTitle("Select Video File")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        if file_dialog.exec_():
            self.video_processor.video_path = file_dialog.selectedFiles()[0]
            self.timer.start(30)  # Update frame every 30 milliseconds
            self.video_processor.start()

    def real_time(self):
        self.video_processor.video_path = 0
        self.timer.start(30)  # Update frame every 30 milliseconds
        self.video_processor.start()

    def update_frame(self):
        # Empty method; processing is handled by the VideoProcessor thread
        pass

    def update_gui(self, img):
        # Display the frame in QLabel
        height, width, channel = img.shape
        bytes_per_line = 3 * width
        q_img = QImage(img.data, width, height, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_img.rgbSwapped())
        self.ui.detection_vidFeedback.setPixmap(pixmap)
        print("Frame Shape:", img.shape)

    def closeEvent(self, event):
        # Release video capture when closing the application
        self.video_processor.cap.release()
        event.accept()

    def stop_detection(self):
        self.video_processor.stop()

    def update_activity(self, index):
        # Get the chosen item from the combo box
        chosen_item = self.ui.comboBox.currentText()

        # Emit the current_activity signal
        self.current_activity.emit(chosen_item)

    def update_time(self):
        # Get the current date and time
        current_datetime = QDateTime.currentDateTime()

        # Format the date and time as a string
        formatted_datetime = current_datetime.toString("yyyy-MM-dd hh:mm:ss")

        # Update the label text
        self.ui.dateTime_label.setText(f"Current Time and Date: {formatted_datetime}")

    def success_message(self):
        # Set the success label text
        self.ui.success_label.setText("Detection Data Saved to Records")

        # Start a timer to hide the success label after 10 seconds
        self.success_timer = QTimer(self)
        self.ui.success_label.setWordWrap(True)
        self.success_timer.timeout.connect(self.clear_success_label)
        self.success_timer.start(10000)  # 10000 milliseconds = 10 seconds

    def clear_success_label(self):
        # Clear the success label text
        self.ui.success_label.clear()
        # Stop the timer
        self.success_timer.stop()

    def load_records(self):
        excel_file_path = '../data/database/test2.xlsx'
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook['Sheet1']

        self.ui.detection_recordsTable.setRowCount(sheet.max_row - 1)
        self.ui.detection_recordsTable.setColumnCount(sheet.max_column)

        list_values = list(sheet.values)
        self.ui.detection_recordsTable.setHorizontalHeaderLabels(list_values[0])

        row_index = 0
        for value_tuple in list_values[1:]:
            col_index = 0
            for value in value_tuple:
                self.ui.detection_recordsTable.setItem(row_index, col_index, QTableWidgetItem(str(value)))
                col_index += 1
            row_index += 1

    def refresh_table(self):
        self.load_records()
    def playAuditoryGif(self):
        self.movie = QMovie('assets/auditoryLearner.gif')
        self.movie.setScaledSize(QSize(210, 200))
        self.ui.auditory_label.setMovie(self.movie)
        self.ui.auditory_label.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playVisualGif(self):
        self.movie = QMovie('assets/visualLearner.gif')
        self.movie.setScaledSize(QSize(210, 200))
        self.ui.visual_label.setMovie(self.movie)
        self.ui.visual_label.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playKinestheticGif(self):
        self.movie = QMovie('assets/kinestheticLearner.gif')
        self.movie.setScaledSize(QSize(210, 200))
        self.ui.kinesthetic_label.setMovie(self.movie)
        self.ui.kinesthetic_label.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def alignCenterPercentage(self):
        self.ui.label_19.setAlignment(Qt.AlignCenter)
        self.ui.label_20.setAlignment(Qt.AlignCenter)
        self.ui.label_21.setAlignment(Qt.AlignCenter)

    def playAuditorySuggestGif1(self):
        self.movie = QMovie('assets/auditoryGif1.gif')
        self.movie.setScaledSize(QSize(230, 200))
        self.ui.audioSuggest_label1.setMovie(self.movie)
        self.ui.audioSuggest_label1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playAuditorySuggestGif2(self):
        self.movie = QMovie('assets/auditoryGif2.gif')
        self.movie.setScaledSize(QSize(230, 190))
        self.ui.audioSuggest_label2.setMovie(self.movie)
        self.ui.audioSuggest_label2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playAuditorySuggestGif3(self):
        self.movie = QMovie('assets/auditoryGif3.gif')
        self.movie.setScaledSize(QSize(200, 180))
        self.ui.audioSuggest_label3.setMovie(self.movie)
        self.ui.audioSuggest_label3.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playAuditorySuggestGif4(self):
        self.movie = QMovie('assets/auditoryGif4.gif')
        self.movie.setScaledSize(QSize(220, 180))
        self.ui.audioSuggest_label4.setMovie(self.movie)
        self.ui.audioSuggest_label4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playVisualSuggestGif1(self):
        self.movie = QMovie('assets/visualGif1.gif')
        self.movie.setScaledSize(QSize(190, 190))
        self.ui.visualSuggest_label1.setMovie(self.movie)
        self.ui.visualSuggest_label1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playVisualSuggestGif2(self):
        self.movie = QMovie('assets/visualGif2.gif')
        self.movie.setScaledSize(QSize(220, 190))
        self.ui.visualSuggest_label2.setMovie(self.movie)
        self.ui.visualSuggest_label2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playVisualSuggestGif3(self):
        self.movie = QMovie('assets/visualGif3.gif')
        self.movie.setScaledSize(QSize(200, 180))
        self.ui.visualSuggest_label3.setMovie(self.movie)
        self.ui.visualSuggest_label3.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playVisualSuggestGif4(self):
        self.movie = QMovie('assets/visualGif4.gif')
        self.movie.setScaledSize(QSize(210, 180))
        self.ui.visualSuggest_label4.setMovie(self.movie)
        self.ui.visualSuggest_label4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playKinestheticSuggestGif1(self):
        self.movie = QMovie('assets/kinestheticGif1.gif')
        self.movie.setScaledSize(QSize(190, 190))
        self.ui.kinestheticSuggest_label1.setMovie(self.movie)
        self.ui.kinestheticSuggest_label1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playKinestheticSuggestGif2(self):
        self.movie = QMovie('assets/kinestheticGif2.gif')
        self.movie.setScaledSize(QSize(220, 190))
        self.ui.kinestheticSuggest_label2.setMovie(self.movie)
        self.ui.kinestheticSuggest_label2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playKinestheticSuggestGif3(self):
        self.movie = QMovie('assets/kinestheticGif3.gif')
        self.movie.setScaledSize(QSize(200, 180))
        self.ui.kinestheticSuggest_label3.setMovie(self.movie)
        self.ui.kinestheticSuggest_label3.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def playKinestheticSuggestGif4(self):
        self.movie = QMovie('assets/kinestheticGif4.gif')
        self.movie.setScaledSize(QSize(210, 180))
        self.ui.kinestheticSuggest_label4.setMovie(self.movie)
        self.ui.kinestheticSuggest_label4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixAuditorySuggestions1(self):
        self.ui.audioSuggest_title1.setWordWrap(True)
        self.ui.audioSuggest_desc1.setWordWrap(True)
        self.movie = QMovie('assets/auditoryGif1.gif')
        self.movie.setScaledSize(QSize(340, 340))
        self.ui.audioSuggest_gifLabel1.setMovie(self.movie)
        self.ui.audioSuggest_gifLabel1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixAuditorySuggestions2(self):
        self.ui.audioSuggest_title2.setWordWrap(True)
        self.ui.audioSuggest_desc2.setWordWrap(True)
        self.movie = QMovie('assets/auditoryGif2.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.audioSuggest_gifLabel2.setMovie(self.movie)
        self.ui.audioSuggest_gifLabel2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixAuditorySuggestions3(self):
        self.ui.audioSuggest_title3.setWordWrap(True)
        self.ui.audioSuggest_desc3.setWordWrap(True)
        self.movie = QMovie('assets/auditoryGif3.gif')
        self.movie.setScaledSize(QSize(340, 340))
        self.ui.audioSuggest_gifLabel3.setMovie(self.movie)
        self.ui.audioSuggest_gifLabel3.setAlignment(Qt.AlignCenter)
        self.ui.activity_label.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixAuditorySuggestions4(self):
        self.ui.audioSuggest_title4.setWordWrap(True)
        self.ui.audioSuggest_desc4.setWordWrap(True)
        self.movie = QMovie('assets/auditoryGif4.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.audioSuggest_gifLabel4.setMovie(self.movie)
        self.ui.audioSuggest_gifLabel4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixVisualSuggestions1(self):
        self.ui.visualSuggest_title1.setWordWrap(True)
        self.ui.visualSuggest_desc1.setWordWrap(True)
        self.movie = QMovie('assets/visualGif1.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.visualSuggest_gifLabel1.setMovie(self.movie)
        self.ui.visualSuggest_gifLabel1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixVisualSuggestions2(self):
        self.ui.visualSuggest_title2.setWordWrap(True)
        self.ui.visualSuggest_desc2.setWordWrap(True)
        self.movie = QMovie('assets/visualGif2.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.visualSuggest_gifLabel2.setMovie(self.movie)
        self.ui.visualSuggest_gifLabel2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixVisualSuggestions3(self):
        self.ui.visualSuggest_title3.setWordWrap(True)
        self.ui.visualSuggest_desc3.setWordWrap(True)
        self.movie = QMovie('assets/visualGif3.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.visualSuggest_gifLabel3.setMovie(self.movie)
        self.ui.visualSuggest_gifLabel3.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixVisualSuggestions4(self):
        self.ui.visualSuggest_title4.setWordWrap(True)
        self.ui.visualSuggest_desc4.setWordWrap(True)
        self.movie = QMovie('assets/visualGif4.gif')
        self.movie.setScaledSize(QSize(340, 320))
        self.ui.visualSuggest_gifLabel4.setMovie(self.movie)
        self.ui.visualSuggest_gifLabel4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixKinestheticSuggestions1(self):
        self.ui.kinestheticSuggest_title1.setWordWrap(True)
        self.ui.kinestheticSuggest_desc1.setWordWrap(True)
        self.movie = QMovie('assets/kinestheticGif1.gif')
        self.movie.setScaledSize(QSize(320, 320))
        self.ui.kinestheticSuggest_gifLabel1.setMovie(self.movie)
        self.ui.kinestheticSuggest_gifLabel1.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixKinestheticSuggestions2(self):
        self.ui.kinestheticSuggest_title2.setWordWrap(True)
        self.ui.kinestheticSuggest_desc2.setWordWrap(True)
        self.movie = QMovie('assets/kinestheticGif2.gif')
        self.movie.setScaledSize(QSize(340, 310))
        self.ui.kinestheticSuggest_gifLabel2.setMovie(self.movie)
        self.ui.kinestheticSuggest_gifLabel2.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixKinestheticSuggestions3(self):
        self.ui.kinestheticSuggest_title3.setWordWrap(True)
        self.ui.kinestheticSuggest_desc3.setWordWrap(True)
        self.movie = QMovie('assets/kinestheticGif3.gif')
        self.movie.setScaledSize(QSize(320, 320))
        self.ui.kinestheticSuggest_gifLabel3.setMovie(self.movie)
        self.ui.kinestheticSuggest_gifLabel3.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def fixKinestheticSuggestions4(self):
        self.ui.kinestheticSuggest_title4.setWordWrap(True)
        self.ui.kinestheticSuggest_desc4.setWordWrap(True)
        self.movie = QMovie('assets/kinestheticGif4.gif')
        self.movie.setScaledSize(QSize(340, 310))
        self.ui.kinestheticSuggest_gifLabel4.setMovie(self.movie)
        self.ui.kinestheticSuggest_gifLabel4.setAlignment(Qt.AlignCenter)
        self.movie.start()

    def create_bar(self):
        set0 = QBarSet("Auditory")
        set1 = QBarSet("Visual")
        set2 = QBarSet("Kinesthetic")

        set0 << 1 << 2 << 3 << 4 << 5 << 6
        set1 << 3 << 2 << 6 << 5 << 5 << 3
        set2 << 2 << 2 << 3 << 4 << 4 << 3


        set0.setBrush(QColor(83, 91, 191))
        set1.setBrush(QColor(122, 130, 233))
        set2.setBrush(QColor(71, 76, 138))


        series = QPercentBarSeries()
        series.append(set0)
        series.append(set1)
        series.append(set2)

        chart = QChart()
        chart.addSeries(series)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QColor(240, 240, 240, 0))  # Set light gray background

        categories = ["Jan", "Feb", "March", "Apr", "May", "Jun"]
        axis = QBarCategoryAxis()
        axis.append(categories)
        chart.createDefaultAxes()
        chart.setAxisX(axis, series)

        chartView = QChartView(chart)
        self.ui.percentage_bar_chart_cont.addWidget(chartView)

    def create_donutchart(self):

        slice0 = QPieSlice("Clapping", 160)
        slice1 = QPieSlice("Listening", 80)
        slice2 = QPieSlice("Tapping", 50)
        slice3 = QPieSlice("Curious", 30)
        slice4 = QPieSlice("Bored", 10)

        slice0.setBrush(QColor(83, 91, 191))
        slice1.setBrush(QColor(122, 130, 233))
        slice2.setBrush(QColor(71, 76, 138))
        slice3.setBrush(QColor(74, 82, 191))
        slice4.setBrush(QColor(140, 82, 255))

        series = QPieSeries()
        series.setHoleSize(0.40)
        series.append(slice0)
        series.append(slice1)
        series.append(slice2)
        series.append(slice3)
        series.append(slice4)

        chart = QChart()
        chart.addSeries(series)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QColor(240, 240, 240, 0))

        chartView = QChartView(chart)
        self.ui.donut_con_chart.addWidget(chartView)

    def create_donutchart1(self):

        slice0 = QPieSlice("Clapping", 70)
        slice1 = QPieSlice("Listening", 120)
        slice2 = QPieSlice("Tapping", 50)
        slice3 = QPieSlice("Curious", 120)
        slice4 = QPieSlice("Bored", 10)

        slice0.setBrush(QColor(83, 91, 191))
        slice1.setBrush(QColor(122, 130, 233))
        slice2.setBrush(QColor(71, 76, 138))
        slice3.setBrush(QColor(74, 82, 191))
        slice4.setBrush(QColor(140, 82, 255))

        series = QPieSeries()
        series.setHoleSize(0.40)
        series.append(slice0)
        series.append(slice1)
        series.append(slice2)
        series.append(slice3)
        series.append(slice4)

        chart = QChart()
        chart.addSeries(series)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QColor(240, 240, 240, 0))

        chartView = QChartView(chart)
        self.ui.donut_con_chart_2.addWidget(chartView)

    def create_donutchart2(self):

        slice0 = QPieSlice("Clapping", 200)
        slice1 = QPieSlice("Listening", 150)
        slice2 = QPieSlice("Tapping", 200)
        slice3 = QPieSlice("Curious", 120)
        slice4 = QPieSlice("Bored", 10)

        slice0.setBrush(QColor(83, 91, 191))
        slice1.setBrush(QColor(122, 130, 233))
        slice2.setBrush(QColor(71, 76, 138))
        slice3.setBrush(QColor(74, 82, 191))
        slice4.setBrush(QColor(140, 82, 255))

        series = QPieSeries()
        series.setHoleSize(0.40)
        series.append(slice0)
        series.append(slice1)
        series.append(slice2)
        series.append(slice3)
        series.append(slice4)

        chart = QChart()
        chart.addSeries(series)
        chart.setAnimationOptions(QChart.SeriesAnimations)
        chart.setBackgroundBrush(QColor(240, 240, 240, 0))

        chartView = QChartView(chart)
        self.ui.donut_con_chart_3.addWidget(chartView)

    def create_linegraph(self):
        chart = QChart()
        # Create data for each line
        data1 = [(1, 2), (2, 4), (3, 3), (4, 5)]
        data2 = [(2, 3), (3, 5), (4, 4), (5, 2)]
        data3 = [(3, 4), (4, 2), (5, 5), (6, 3)]

        # Create QLineSeries objects
        series1 = QLineSeries()
        series1.setName("Clapping")
        series1.setColor(QColor(83, 91, 191))
        for data_point in data1:
            x, y = data_point
            point = QtCore.QPointF(x, y)
            series1.append(point)

        series2 = QLineSeries()
        series2.setName("Curious")
        series2.setColor(QColor(122, 130, 233))
        for data_point in data2:
            x, y = data_point
            point = QtCore.QPointF(x, y)
            series2.append(point)

        series3 = QLineSeries()
        series3.setName("Listening")
        series3.setColor(QColor(71, 76, 138))
        for data_point in data3:
            x, y = data_point
            point = QtCore.QPointF(x, y)
            series3.append(point)

        chart.addSeries(series1)
        chart.addSeries(series2)
        chart.addSeries(series3)

        chart.setAnimationOptions(QChart.AllAnimations)
        chart.setBackgroundBrush(QColor(240, 240, 240, 0))

        chart.createDefaultAxes()

        chartView = QChartView(chart)
        self.ui.line_chart_cont.addWidget(chartView)


    def showDetection(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.detection)
    def showBarGraph(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.bar_graph)
    def showNestedDonuts(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.nested_donuts)
    def showNestedDonuts1(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.nested_donuts1)

    def showNestedDonuts2(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.nested_donuts2)
    def showLineChart(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.line_chart)
    def showHybridGraph(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.hybrid_graph)
    def showPrediction(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.prediction)
    def showAuditorySuggestions(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.audio_suggest)

    def showAuditorySuggestions1(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.audio_suggest1)

    def showAuditorySuggestions2(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.audio_suggest2)

    def showAuditorySuggestions3(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.audio_suggest3)

    def showAuditorySuggestions4(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.audio_suggest4)

    def showVisualSuggestions(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.visual_suggest)

    def showVisualSuggestions1(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.visual_suggest1)

    def showVisualSuggestions2(self):

        self.ui.stackedWidget.setCurrentWidget(self.ui.visual_suggest2)
    def showVisualSuggestions3(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.visual_suggest3)

    def showVisualSuggestions4(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.visual_suggest4)

    def showKinestheticSuggestions(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.kinesthetic_suggest)

    def showKinestheticSuggestions1(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.kinesthetic_suggest1)

    def showKinestheticSuggestions2(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.kinesthetic_suggest2)

    def showKinestheticSuggestions3(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.kinesthetic_suggest3)

    def showKinestheticSuggestions4(self):
        self.ui.stackedWidget.setCurrentWidget(self.ui.kinesthetic_suggest4)

# Execute App
def main():
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
# End


